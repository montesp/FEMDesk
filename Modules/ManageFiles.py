from logging.config import valid_ident
import opcode
import os
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QFileDialog, QWidget, QLineEdit, QMessageBox
from PyQt5.QtCore import Qt, QPointF
from PyQt5.QtGui import QBrush, QPolygonF
from pyparsing import col
from Modules.Dictionary.DMatrix import *
from Modules.Dictionary.DFiles import *
from Modules.Dictionary.DModelWizard import *
from Modules.Matrix import *
import Modules.ModelWizard
import Modules.Materials
import numpy as np
import Modules.Tabs
from Modules.SaveExcel import *
from Modules.LoadExcel import *



class openSaveDialog(QWidget):
    def __init__(self):
        super().__init__()
        self.windowTitle("Ingrese el nombre")

class wbSheet(object):
    def __init__(self, wb, wb1, wb2, wb3, wb4, wb5, wb6, wb7, wb8, wbPolygons, wbMaterials):
        self.wb1 = wb1
        self.wb2 = wb2
        self.wb3 = wb3
        self.wb4 = wb4
        self.wb5 = wb5
        self.wb6 = wb6
        self.wb7 = wb7
        self.wb8 = wb8
        self.wbMaterials = wbMaterials                                          
        self.wbPolygons = wbPolygons
class FileData():
    #Función para buscar en el explorador un archivo excel para abrir la configuracion guardada y usarla en el programa
    def getFileName(self, material, canvas):
        option = QFileDialog.Option()
        file_filter= 'Excel File (*.xlsx *.xls)'
        file = QFileDialog.getOpenFileName(
            self,
            caption='Select a file',
            directory= "Saves",
            filter=file_filter,
            initialFilter='Excel File (*.xlsx *.xls)',
            options=option
        )
        if file != '':
          #try:
                wb = load_workbook(file[0])
                sheet = wb.active
                FileData.loadData(self, sheet, wb, material, canvas)
                directory["dir"] = str(file[0])
                self.lblDirectory.setText(directory["dir"])
                print(directory)
          #except Exception:
                #print("Operacion Cancelada")

    #Función para decirle al indicador si la configuración del programa fue modificada
    #Esto solo en caso de que se encuentre un archivo excel cargado
    def checkUpdateFile(self):
        fileIndicator["*"] = "*"
        if directory["dir"] != '':
            self.lblDirectory.setText(directory["dir"] + fileIndicator["*"])
            self.actionSaves.setEnabled(True)

    #Funcion para decirle al indicador si la configuracion del programa no tiene 
    #modificaciones
    def uncheckUpdateFile(self):
        fileIndicator["*"] = ""
        self.lblDirectory.setText(directory["dir"] + fileIndicator["*"])
        self.actionSaves.setEnabled(False)
        
    def newFileName(self, material, canvas):
        wb = Workbook()
        sheet = wb.active

        option = QFileDialog.Options()
        file = QFileDialog.getSaveFileName(QWidget(), 
        "Save File", 
        "Saves\\.xlsx", 
        "Excel File (*.xlsx *.xls)", 
        options=option)

        if file != '':
          #try:
                fileName = file[0]
                FileData.newData(self, fileName, wb, sheet, material, canvas)
                directory["dir"] = str(file[0])
                self.lblDirectory.setText(directory["dir"])
          #except Exception:
                #print("Operacion Cancelada")

    def saveAsFile(self, material, canvas):
        wb = Workbook()
        sheet = wb.active

        option = QFileDialog.Options()
        file = QFileDialog.getSaveFileName(QWidget(), 
        "Save File", 
        "Saves\\.xlsx", 
        "Excel File (*.xlsx *.xls)", 
        options=option)

        if file != '':
          try:
                fileName = file[0]
                FileData.newData(self, fileName, wb, sheet, material, canvas)
                directory["dir"] = str(file[0])
                self.lblDirectory.setText(directory["dir"])
          except Exception:
                print("Operacion Cancelada")

    #Función mandada a llamar desde ActionSaves, se encarga de conectar con el archivo excel y actializarlo con nuevos datos
    def updateFile(self, material, canvas):
        wb = load_workbook(directory["dir"])
        sheet = wb.active
        
        file = directory["dir"]
        
        wb1 = wb["diffusion"]
        wb2 = wb["absorption"]
        wb3 = wb["source"]
        wb4 = wb["mass"]
        wb5 = wb["damMass"]
        wb6 = wb["cFlux"]
        wb7 = wb["convection"]
        wb8 = wb["cSource"]
        wbPolygons= wb["polygons"]
        wbMaterials = wb["materials"]

        wbSheet = Modules.ManageFiles.wbSheet(self, wb1, wb2, wb3, wb4, wb5, wb6, wb7, wb8, wbPolygons, wbMaterials)

        FileData.newWriteData(self, file, wb, sheet, wbSheet, material, canvas)
        
        
    def resetFile(self, material, canvas):
        # print(fileIndicator["*"])
        if fileIndicator["*"] == '*':
         dialog = QMessageBox()
         dialog.setWindowTitle('Aviso')
         dialog.setText('¿Seguro que quieres cerrar el archivo? Se borrarán los cambios no guardados')
         yesdialog = dialog.addButton('Si', QMessageBox.YesRole)
         nodialog = dialog.addButton('No', QMessageBox.NoRole)
         savedialog = dialog.addButton('Guardar Cambios', QMessageBox.YesRole)
         dialog.exec_()
         if dialog.clickedButton() == yesdialog: 
                FileData.resetData(self)
         elif dialog.clickedButton() == nodialog:
                print("Operación Cancelada")
         elif dialog.clickedButton() == savedialog:
                FileData.updateFile(self, material, canvas)
                FileData.resetData(self)
        else:
                FileData.resetData(self)

    
       
    #Funcion para configurar el archivo EXCEL de modo que puede ser usado
    #para guardar los datos del programa
    def newData(self, file, wb, sheet, material, canvas):
        #Crear las paginas del archivo Excel
        wb1 = wb.create_sheet('diffusion')
        wb2 = wb.create_sheet('absorption')
        wb3 = wb.create_sheet('source')
        wb4 = wb.create_sheet('mass')
        wb5 = wb.create_sheet('damMass')
        wb6 = wb.create_sheet('cFlux')
        wb7 = wb.create_sheet('convection')
        wb8 = wb.create_sheet('cSource')
        wbMaterials = wb.create_sheet('materials')                                          
        wbPolygons = wb.create_sheet('polygons')
        #Mandar a llamar la funcion para guardar las paginas del archivo Excel
        wbSheet = Modules.ManageFiles.wbSheet(self, wb1, wb2, wb3, wb4, wb5, wb6, wb7, wb8, wbPolygons, wbMaterials)

        #Ajustar las dimensiones de las columnas en el Excel
        SaveExcel.adjustExcelDimensions(self, sheet)
        
        #Escribir los labels en el archivo Excel
        SaveExcel.writeExcelText(self, sheet, wbSheet)
       
        #Llamar la funcion para guardar los datos en el archivo excel
        FileData.newWriteData(self, file, wb, sheet, wbSheet, material, canvas)
        
   
    #Guardar los datos del programa al archivo Excel
    def newWriteData(self, file, wb, sheet, wbSheet, material, canvas):
        #Guardar los datos de los items de Coefficient PDE en el Excel
        SaveExcel.saveExcelItemsPDE(self, sheet)
        #Guardar los datos de las coordenadas de los QComboBox en el Excel
        SaveExcel.saveExcelCoordinates(self, sheet)        
        #Guardar los datos de la clase Materials en el archivo Excel
        SaveExcel.saveExcelMaterialsData(self, wbSheet, material)
        #Guardar los datos de las matrices en el archivo Excel
        SaveExcel.saveExcelMatrixData(self, wbSheet)
        #Guardar los datos de las figuras en el archivo Excel
        SaveExcel.saveExcelFigures(self, wbSheet, canvas)

        wb.save(file)
        FileData.uncheckUpdateFile(self)
        self.actionSave_As.setEnabled(True)
        self.actionClose.setEnabled(True)
        QMessageBox.information(self, "Important message", "Guardado Exitoso")
        

    
    #Función para cargar la configuración
    def loadData(self, sheet, wb, material, canvas):
        #Cargar las paginas del archivo Excel
        wb1 = wb["diffusion"]
        wb2 = wb["absorption"]
        wb3 = wb["source"]
        wb4 = wb["mass"]
        wb5 = wb["damMass"]
        wb6 = wb["cFlux"]
        wb7 = wb["convection"]
        wb8 = wb["cSource"]
        wbMaterials = wb["materials"]
        wbPolygons = wb["polygons"]
        wbSheet = Modules.ManageFiles.wbSheet(self, wb1, wb2, wb3, wb4, wb5, wb6, wb7, wb8, wbPolygons, wbMaterials)

        #Cargar las dimensiones de las matrices 
        LoadExcel.loadExcelMatrixDimensions(self, sheet)
        #Cargar los datos de los items del Coefficient PDE
        LoadExcel.loadExcelItemsData(self, sheet)    
        #Cargar las coordenadas de los QCombobox del Coefficient PDE
        LoadExcel.loadExcelCoordinates(self, sheet)        
        #Cargar la lista de items activados del Coefficient PDE
        LoadExcel.loadExcelItemsCoefficientPDE(self)
        #Cargar los datos de las matrices
        LoadExcel.loadExcelMatrixData(self, wbSheet)
        #Cargar los antiguos datos mostrados de cada coordenada del combobox
        LoadExcel.loadExcelCoordinateData(self)
        #Cargar los datos de la clase materials del archivo Excel
        LoadExcel.loadExcelMaterialsData(self, wbSheet, material)
        #Cargar la configuracion mas reciente del ModelWizard
        LoadExcel.loadExcelModelWizard(self)
        #Cargar las figuras guardadas en el archivo Excel
        LoadExcel.loadExcelFigures(self, wbSheet, canvas)
       
        #Decirle al programa que no hay edicione sen el archivo actual
        FileData.uncheckUpdateFile(self)
        self.actionSave_As.setEnabled(True)
        self.actionClose.setEnabled(True)


 
    def resetData(self):
        #Ocultar todos los items del ToolBox Coefficients PDE y dejar solo el item Initial Values
        for i in range(1, self.CoefficentForM.count()):
            self.CoefficentForM.removeItem(1)

        for i, item in enumerate(self.CoefficientCheckBoxArray):
                self.CoefficientCheckBoxArray[i - 1].setChecked(False)

        #Resetear todos los combobox de las secciones y dejarles solo el valor de 1
        for i, item in enumerate(self.arrayCmbRowColumns):
         for j, item in enumerate(self.arrayCmbRowColumns[i]):
                        self.arrayCmbRowColumns[i][j].clear()
                        self.arrayCmbRowColumns[i][j].addItem("1")

        #Limpiar todos los lineEdits de cada seccion
        for i, item in enumerate(self.arraylEditsCoefficientsPDE):
         for j, item in enumerate(self.arraylEditsCoefficientsPDE[i]):
                        self.arraylEditsCoefficientsPDE[i][j].setText("")

        #Resetear el combobox de Initial Values
        self.cmbInitialValues.clear()
        self.cmbInitialValues.addItem("u1")
        #Eliminar la dirección del archivo excel actual del QLabel
        self.lblDirectory.setText("")

        self.actionSaves.setEnabled(False)
        self.actionSave_As.setEnabled(False)
        self.actionClose.setEnabled(False)

        #Resetear el modo de input del diffusion matrix a 1 (isotrópico)
        diffusionMatrix["inputMode"] = 0
        self.cmbDiffusionCoef.setCurrentIndex(diffusionMatrix["inputMode"])
        noItemsCoeffM["noItems"] = 0
        noItemsCoeffM["items"] = 0
        initialValues["noVariables"] = 1
        self.inputDepedentVarial.setText(str(initialValues["noVariables"]))
        fileIndicator["*"] = ""
        #Eliminar la dirección del archivo excel en la memoria de la variable
        directory["dir"] = ""
        
        #Resetear la configuracion del ModelWizard
        myFlags["ModelWizardMode"] = "None"
        self.itemSpace[0].setExpanded(False)
        self.item2D[0].setExpanded(False)
        self.itemPhysics[0].setExpanded(False)
        self.itemHeat[0].setExpanded(False)
        self.itemMath[0].setExpanded(False)
        self.itemFluids[0].setForeground(0, QBrush(Qt.black))
        self.itemPDE[0].setForeground(0, QBrush(Qt.black))
        self.itemSolids[0].setForeground(0, QBrush(Qt.black))
        Modules.ModelWizard.ModelWizard.flagModelWizardActivated = False
        Modules.Tabs.Tabs.hideElementsTab(self.tabs, self.tabWidgetMenu)

    def resetDataWithoutLoseFile(self):
        #Ocultar todos los items del ToolBox Coefficients PDE y dejar solo el item Initial Values
        for i in range(1, self.CoefficentForM.count()):
            self.CoefficentForM.removeItem(1)

        for i, item in enumerate(self.CoefficientCheckBoxArray):
                self.CoefficientCheckBoxArray[i - 1].setChecked(False)

        #Resetear todos los combobox de las secciones y dejarles solo el valor de 1
        for i, item in enumerate(self.arrayCmbRowColumns):
         for j, item in enumerate(self.arrayCmbRowColumns[i]):
                        self.arrayCmbRowColumns[i][j].clear()
                        self.arrayCmbRowColumns[i][j].addItem("1")

        #Limpiar todos los lineEdits de cada seccion
        for i, item in enumerate(self.arraylEditsCoefficientsPDE):
         for j, item in enumerate(self.arraylEditsCoefficientsPDE[i]):
                        self.arraylEditsCoefficientsPDE[i][j].setText("")

        #Resetear el combobox de Initial Values
        self.cmbInitialValues.clear()
        self.cmbInitialValues.addItem("u1")
        #Eliminar la dirección del archivo excel actual del QLabel
        self.lblDirectory.setText("")

        if directory["dir"] != "":
         self.actionSaves.setEnabled(True)
         self.actionSave_As.setEnabled(True)
         self.actionClose.setEnabled(True)

        #Resetear el modo de input del diffusion matrix a 1 (isotrópico)
        diffusionMatrix["inputMode"] = 0
        self.cmbDiffusionCoef.setCurrentIndex(diffusionMatrix["inputMode"])
        noItemsCoeffM["noItems"] = 0
        noItemsCoeffM["items"] = 0
        initialValues["noVariables"] = 1
        self.inputDepedentVarial.setText(str(initialValues["noVariables"]))
        fileIndicator["*"] = ""
        
        myFlags["ModelWizardMode"] = "None"

        Modules.ModelWizard.ModelWizard.flagModelWizardActivated = False
        Modules.Tabs.Tabs.hideElementsTab(self.tabs, self.tabWidgetMenu)
    
        
class Update():
 #Funcion para actualizar los datos segun las coordenadas mas recientes
 def currentData(self, pos):
        if pos == 1:
            coordinates["coordinateDiffusion"] = [self.cmbRowDiffusionCoef.currentIndex(), self.cmbColumnDiffusionCoef.currentIndex()]
            positionMatrix = allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()]
            if positionMatrix == 'None' or positionMatrix == '':
                floatMatrix = 0
            else:
                positionMatrix = positionMatrix.replace(" ","")
                positionMatrix = positionMatrix.strip('[]')
                positionMatrix = positionMatrix.split(',')
                floatMatrix = ['{0:g}'.format(float(i)) for i in positionMatrix]
                floatMatrix = int(floatMatrix[4])

            if floatMatrix == 0:
             self.cmbDiffusionCoef.setCurrentIndex(floatMatrix)   
             if allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()] == 'None' or allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()] == '':
              self.lEditDiffusionCoef.setText("")
             else: 
              strCell = allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()]
              strCell = strCell.strip("[]")
              strCell = strCell.split(',')
              floatCell = ['{0:g}'.format(float(i)) for i in strCell]
              self.lEditDiffusionCoef.setText(floatCell[0])
         
            if floatMatrix == 1:
             self.cmbDiffusionCoef.setCurrentIndex(floatMatrix)   
             if allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()] == 'None' or allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()] == '':
              self.lEditDiffusionCoef11.setText("")
              self.lEditDiffusionCoef22.setText("")
             else:  
              strCell = allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()]
              strCell = strCell.strip("[]")
              strCell = strCell.split(',')
              floatCell = ['{0:g}'.format(float(i)) for i in strCell]
              self.lEditDiffusionCoef11.setText(floatCell[0])
              self.lEditDiffusionCoef22.setText(floatCell[3])
           
            if floatMatrix == 2 or floatMatrix == 3:
              self.cmbDiffusionCoef.setCurrentIndex(floatMatrix)  
              if allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()] == 'None' or allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()] == '':
                self.lEditDiffusionCoef11.setText("")
                self.lEditDiffusionCoef12.setText("")
                self.lEditDiffusionCoef21.setText("")
                self.lEditDiffusionCoef22.setText("")
              else:
                strCell = allNewMatrix.diffusionM[self.cmbRowDiffusionCoef.currentIndex()][self.cmbColumnDiffusionCoef.currentIndex()]
                strCell = strCell.strip("[]")
                strCell = strCell.split(',')
                floatCell = ['{0:g}'.format(float(i)) for i in strCell]
                self.lEditDiffusionCoef11.setText(floatCell[0])
                self.lEditDiffusionCoef12.setText(floatCell[1])
                self.lEditDiffusionCoef21.setText(floatCell[2])
                self.lEditDiffusionCoef22.setText(floatCell[3])

        if pos == 2:
            coordinates["coordinateAbsorption"] = [self.cmbAbsorptionRow.currentIndex(), self.cmbAbsorptionColumn.currentIndex()]
            if allNewMatrix.absorptionM[self.cmbAbsorptionRow.currentIndex()][self.cmbAbsorptionColumn.currentIndex()] == 'None' or allNewMatrix.absorptionM[self.cmbAbsorptionRow.currentIndex()][self.cmbAbsorptionColumn.currentIndex()] == '':    
             self.lEditAbsorCoef.setText("")
            else:
             self.lEditAbsorCoef.setText(allNewMatrix.absorptionM[self.cmbAbsorptionRow.currentIndex()][self.cmbAbsorptionColumn.currentIndex()])
        if pos == 3:
            coordinates["coordinateSource"] = [self.cmbSourceRow.currentIndex()]
            if allNewMatrix.sourceM[self.cmbSourceRow.currentIndex()] == 'None' or allNewMatrix.sourceM[self.cmbSourceRow.currentIndex()] == '':
             self.lEditSourceTerm.setText("")
            else:
             self.lEditSourceTerm.setText(allNewMatrix.sourceM[self.cmbSourceRow.currentIndex()])
        if pos == 4:
            coordinates["coordinateMass"] = [self.cmbMassCoefRow.currentIndex(), self.cmbMassCoefColumn.currentIndex()]
            if allNewMatrix.massM[self.cmbMassCoefRow.currentIndex()][self.cmbMassCoefColumn.currentIndex()] == 'None' or allNewMatrix.massM[self.cmbMassCoefRow.currentIndex()][self.cmbMassCoefColumn.currentIndex()] == '':
             self.lEditMassCoef.setText("")
            else:
             self.lEditMassCoef.setText(allNewMatrix.massM[self.cmbMassCoefRow.currentIndex()][self.cmbMassCoefColumn.currentIndex()])
        if pos == 5:
            coordinates["coordinateDamMass"] = [self.cmbDamMassCoefRow.currentIndex(), self.cmbDamMassCoefColumn.currentIndex()]
            if allNewMatrix.damMassM[self.cmbDamMassCoefRow.currentIndex()][self.cmbDamMassCoefColumn.currentIndex()] == 'None' or allNewMatrix.damMassM[self.cmbDamMassCoefRow.currentIndex()][self.cmbDamMassCoefColumn.currentIndex()] == '':
             self.lEditDamMassCoef.setText("")
            else:
             self.lEditDamMassCoef.setText(allNewMatrix.damMassM[self.cmbDamMassCoefRow.currentIndex()][self.cmbDamMassCoefColumn.currentIndex()])
        if pos == 6:
            coordinates["coordinateCFlux"] = [self.cmbCFluxRow.currentIndex(), self.cmbCFluxColumn.currentIndex()]
            if allNewMatrix.cFluxM[self.cmbCFluxRow.currentIndex()][self.cmbCFluxColumn.currentIndex()] == 'None' or allNewMatrix.cFluxM[self.cmbCFluxRow.currentIndex()][self.cmbCFluxColumn.currentIndex()] == '':
                self.lEditAlphaXCFlux.setText("")
                self.lEditAlphaCYFlux.setText("")
            else:
                strCell = allNewMatrix.cFluxM[self.cmbCFluxRow.currentIndex(), self.cmbCFluxColumn.currentIndex()]
                strCell = strCell.strip("[]")
                strCell = strCell.split(',')
                self.lEditAlphaXCFlux.setText(strCell[0])
                self.lEditAlphaCYFlux.setText(strCell[1])
        if pos == 7:
            coordinates["coordinateConvection"] = [self.cmbConvectionRow.currentIndex(), self.cmbConvectionColumn.currentIndex()]
            if allNewMatrix.convectionM[self.cmbConvectionRow.currentIndex()][self.cmbConvectionColumn.currentIndex()] == 'None' or allNewMatrix.convectionM[self.cmbConvectionRow.currentIndex()][self.cmbConvectionColumn.currentIndex()] == '':
                self.lEditBetaXConvCoef.setText("")
                self.lEditBetaYConvCoef.setText("")
            else:
                strCell = allNewMatrix.convectionM[self.cmbConvectionRow.currentIndex(), self.cmbConvectionColumn.currentIndex()]
                strCell = strCell.strip("[]")
                strCell = strCell.split(',')
                self.lEditBetaXConvCoef.setText(strCell[0])
                self.lEditBetaYConvCoef.setText(strCell[1])
        if pos == 8: 
            coordinates["coordinateCSource"] = [self.cmbCSourceRow.currentIndex()]
            if allNewMatrix.cSourceM[self.cmbCSourceRow.currentIndex()] == 'None' or allNewMatrix.cSourceM[self.cmbCSourceRow.currentIndex()] == '':
                self.lEditGammaXCFluxSource.setText("")
                self.lEditGammaYCFluxSource.setText("") 
            else:
                strCell = allNewMatrix.cSourceM[self.cmbCSourceRow.currentIndex()]
                strCell = strCell.strip("[]")
                strCell = strCell.split(',')
                self.lEditGammaXCFluxSource.setText(strCell[0])
                self.lEditGammaYCFluxSource.setText(strCell[1])

#Funcion para actualizar la confiracion de los Combobox del Coefficient PDE
 def currentCoordinateMatrix(self, arrayComb):
        strComb = coordinates["coordinateDiffusion"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada Diffusion")
        print(intComb)
        arrayComb[0][0].setCurrentIndex(intComb[0])
        arrayComb[0][1].setCurrentIndex(intComb[1])

        strComb = coordinates["coordinateAbsorption"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada Absorption")
        print(intComb)
        arrayComb[1][0].setCurrentIndex(intComb[0])
        arrayComb[1][1].setCurrentIndex(intComb[1])

        strComb = coordinates["coordinateSource"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada Source")
        print(intComb)
        arrayComb[2][0].setCurrentIndex(intComb[0])

        strComb = coordinates["coordinateMass"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada Mass")
        print(intComb)
        arrayComb[3][0].setCurrentIndex(intComb[0])
        arrayComb[3][1].setCurrentIndex(intComb[1])

        strComb = coordinates["coordinateDamMass"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada DamMass")
        print(intComb)
        arrayComb[4][0].setCurrentIndex(intComb[0])
        arrayComb[4][1].setCurrentIndex(intComb[1])

        strComb = coordinates["coordinateCFlux"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada CFlux")
        print(intComb)
        arrayComb[5][0].setCurrentIndex(intComb[0])
        arrayComb[5][1].setCurrentIndex(intComb[1])

        strComb = coordinates["coordinateConvection"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada Convection")
        print(intComb)
        arrayComb[6][0].setCurrentIndex(intComb[0])
        arrayComb[6][1].setCurrentIndex(intComb[1])

        strComb = coordinates["coordinateCSource"]
        strComb = strComb.strip("[]")
        strComb = strComb.split(',')
        intComb = [int(i) for i in strComb]
        print("Cordenada CSource")
        print(intComb)
        arrayComb[7][0].setCurrentIndex(intComb[0])

        
        







        

        