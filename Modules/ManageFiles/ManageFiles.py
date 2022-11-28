import os
from logging.config import valid_ident

import numpy as np
import opcode
from openpyxl import Workbook, load_workbook
from pyparsing import col
from PyQt5.QtCore import QPointF, Qt
from PyQt5.QtGui import QBrush, QPolygonF
from PyQt5.QtWidgets import QFileDialog, QLineEdit, QMessageBox, QWidget

import Modules.Materials
import Modules.ModelWizard
import Modules.Tabs
from Modules.Dictionary.DFiles import *
from Modules.Dictionary.DMatrix import *
from Modules.Dictionary.DModelWizard import *
from Modules.ManageFiles.LoadExcel import *
from Modules.ManageFiles.Reset import *
from Modules.ManageFiles.SaveExcel import *
from Modules.ManageFiles.Update import *


class openSaveDialog(QWidget):
    def __init__(self):
        super().__init__()
        self.windowTitle("Ingrese el nombre")

class wbSheet(object):
    def __init__(self, sheet, wb1, wb2, wb3, wb4, wb5, wb6, wb7, wb8, wbConditionsPDE, wbConditionsPDEItems,
    wbConditions, wbMatrixItems, wbPolygons, wbMaterials):
        self.sheet = sheet
        self.wb1 = wb1
        self.wb2 = wb2
        self.wb3 = wb3
        self.wb4 = wb4
        self.wb5 = wb5
        self.wb6 = wb6
        self.wb7 = wb7
        self.wb8 = wb8
        self.wbConditionsPDE = wbConditionsPDE
        self.wbConditionsPDEItems = wbConditionsPDEItems
        self.wbConditions = wbConditions      
        self.wbMatrixItems = wbMatrixItems
        self.wbMaterials = wbMaterials                                          
        self.wbPolygons = wbPolygons
class FileData():
    #Función para decirle al indicador si la configuración del programa fue modificada
    #Esto solo en caso de que se encuentre un archivo excel cargado
    def checkUpdateFile(self):
        fileIndicator["*"] = "*"
        if directory["dir"] != '':
            self.lblDirectory.setText(directory["dir"] + fileIndicator["*"])
            self.actionSaves.setEnabled(True)

    #Funcion para decirle al indicador si la configuracion del programa no tiene 
    #modificaciones
    def uncheckUpdateFile(self):
        fileIndicator["*"] = ""
        self.lblDirectory.setText(directory["dir"] + fileIndicator["*"])
        self.actionSaves.setEnabled(False)

    #Función para buscar en el explorador un archivo excel para abrir la configuracion guardada y usarla en el programa
    def getFileName(self, material, canvas, condition, tabs, win):
        option = QFileDialog.Option()
        file_filter= 'Excel File (*.xlsx *.xls)'
        file = QFileDialog.getOpenFileName(
            self,
            caption='Select a file',
            directory= "Saves",
            filter=file_filter,
            initialFilter='Excel File (*.xlsx *.xls)',
            options=option
        )
        if file != '':
          #try:
                wb = load_workbook(file[0])
                sheet = wb.active
                FileData.loadData(self, sheet, wb, material, canvas, condition, tabs, win)
                directory["dir"] = str(file[0])
                self.lblDirectory.setText(directory["dir"])
                print(directory)
          #except Exception:
                #print("Operacion Cancelada")

    #Funcion para crear un nuevo archivo Excel y guardar la informacion    
    def newFileName(self, material, canvas, conditions):
        wb = Workbook()
        sheet = wb.active

        option = QFileDialog.Options()
        file = QFileDialog.getSaveFileName(QWidget(), 
        "Save File", 
        "Saves\\.xlsx", 
        "Excel File (*.xlsx *.xls)", 
        options=option)

        if file != '':
          #try:
                fileName = file[0]
                FileData.newData(self, fileName, wb, sheet, material, canvas, conditions)
                directory["dir"] = str(file[0])
                self.lblDirectory.setText(directory["dir"])
          #except Exception:
                #print("Operacion Cancelada")

    #Funcion para guardar los datos actuales en un archivo Excel diferente
    def saveAsFile(self, material, canvas):
        wb = Workbook()
        sheet = wb.active

        option = QFileDialog.Options()
        file = QFileDialog.getSaveFileName(QWidget(), 
        "Save File", 
        "Saves\\.xlsx", 
        "Excel File (*.xlsx *.xls)", 
        options=option)

        if file != '':
          try:
                fileName = file[0]
                FileData.newData(self, fileName, wb, sheet, material, canvas)
                directory["dir"] = str(file[0])
                self.lblDirectory.setText(directory["dir"])
          except Exception:
                print("Operacion Cancelada")

    #Función mandada a llamar desde ActionSaves, se encarga de conectar con el archivo excel y actializarlo con nuevos datos
    def updateFile(self, material, canvas):
        wb = load_workbook(directory["dir"])
        sheet = wb.active
        
        file = directory["dir"]
        
        wb1 = wb["diffusion"]
        wb2 = wb["absorption"]
        wb3 = wb["source"]
        wb4 = wb["mass"]
        wb5 = wb["damMass"]
        wb6 = wb["cFlux"]
        wb7 = wb["convection"]
        wb8 = wb["cSource"]
        wbPolygons= wb["polygons"]
        wbMaterials = wb["materials"]

        wbSheet = Modules.ManageFiles.ManageFiles.wbSheet(self, wb1, wb2, wb3, wb4, wb5, wb6, wb7, wb8, wbPolygons, wbMaterials)

        FileData.newWriteData(self, file, wb, sheet, wbSheet, material, canvas)
        
    #Funcion para resetear todo el programa
    def resetFile(self, material, canvas):
        # print(fileIndicator["*"])
        if fileIndicator["*"] == '*':
         dialog = QMessageBox()
         dialog.setWindowTitle('Aviso')
         dialog.setText('¿Seguro que quieres cerrar el archivo? Se borrarán los cambios no guardados')
         yesdialog = dialog.addButton('Si', QMessageBox.YesRole)
         nodialog = dialog.addButton('No', QMessageBox.NoRole)
         savedialog = dialog.addButton('Guardar Cambios', QMessageBox.YesRole)
         dialog.exec_()
         if dialog.clickedButton() == yesdialog: 
                FileData.resetData(self, material, canvas)
         elif dialog.clickedButton() == nodialog:
                print("Operación Cancelada")
         elif dialog.clickedButton() == savedialog:
                FileData.updateFile(self, material, canvas)
                FileData.resetData(self, material, canvas)
        else:
                FileData.resetData(self, material, canvas)

    
    #Funcion para configurar el archivo EXCEL de modo que puede ser usado
    #para guardar los datos del programa
    def newData(self, file, wb, sheet, material, canvas, conditions):
        #Crear las paginas del archivo Excel
        wb1 = wb.create_sheet('diffusion')
        wb2 = wb.create_sheet('absorption')
        wb3 = wb.create_sheet('source')
        wb4 = wb.create_sheet('mass')
        wb5 = wb.create_sheet('damMass')
        wb6 = wb.create_sheet('cFlux')
        wb7 = wb.create_sheet('convection')
        wb8 = wb.create_sheet('cSource')
        wbMatrixItems = wb.create_sheet('matrixItems')
        wbConditionsPDE = wb.create_sheet('ConditionsPDE')
        wbConditionsPDEItems = wb.create_sheet('ConditionsPDEItems')
        wbConditions = wb.create_sheet('Conditions')
        wbMaterials = wb.create_sheet('materials')                                          
        wbPolygons = wb.create_sheet('polygons')
        #Mandar a llamar la funcion para guardar las paginas del archivo Excel
        wbSheet = Modules.ManageFiles.ManageFiles.wbSheet(sheet, wb1, wb2, wb3, wb4, wb5, wb6, wb7, 
        wb8, wbConditionsPDE, wbConditionsPDEItems, wbConditions, wbMatrixItems, wbPolygons, wbMaterials)
        #Ajustar las dimensiones de las columnas en el Excel
        SaveExcel.adjustExcelDimensions(self, sheet)
        #Escribir los labels en el archivo Excel
        SaveExcel.writeExcelText(self, sheet, wbSheet)
        #Llamar la funcion para guardar los datos en el archivo excel
        FileData.newWriteData(self, file, wb, sheet, wbSheet, material, canvas, conditions)
        
   
    #Guardar los datos del programa al archivo Excel
    def newWriteData(self, file, wb, sheet, wbSheet, material, canvas, conditions):
        #Guardar los datos de los items de Coefficient PDE en el Excel
        SaveExcel.saveExcelItemsPDE(self, sheet)
        #Guardar los datos de las coordenadas de los QComboBox en el Excel
        SaveExcel.saveExcelCoordinates(self, sheet)        
        #Guardar los datos de la clase Materials en el archivo Excel
        SaveExcel.saveExcelMaterialsData(self, wbSheet, material)
        #Guardar los datos de las matrices en el archivo Excel
        SaveExcel.saveExcelMatrixData(self, wbSheet)
        #Guardar los datos de los items activados en el archivo Excel
        SaveExcel.saveExcelMatrixItems(self, wbSheet)
        #Guardar los datos de la matrices Conditions PDE en el archivo Excel
        SaveExcel.saveExcelConditionsPDE(self, wbSheet)
        #Guardar los datos de los items activados del Conditions PDE en el archivo Excel
        SaveExcel.saveExcelItemsConditions(self, wbSheet)
        #Guardar los datos de Conditions en el archivo Excel
        SaveExcel.saveExcelConditionsData(self, wbSheet, conditions)
        #Guardar los datos de las figuras en el archivo Excel
        SaveExcel.saveExcelFigures(self, wbSheet, canvas)

        wb.save(file)
        FileData.uncheckUpdateFile(self)
        self.actionSave_As.setEnabled(True)
        self.actionClose.setEnabled(True)
        QMessageBox.information(self, "Important message", "Guardado Exitoso")
        
    #Función para cargar la configuración
    def loadData(self, sheet, wb, material, canvas, condition, tabs, win):
        #Cargar las paginas del archivo Excel
        sheet = wb['Sheet']
        wb1 = wb["diffusion"]
        wb2 = wb["absorption"]
        wb3 = wb["source"]
        wb4 = wb["mass"]
        wb5 = wb["damMass"]
        wb6 = wb["cFlux"]
        wb7 = wb["convection"]
        wb8 = wb["cSource"]
        wbMatrixItems = wb['matrixItems']
        wbConditionsPDE = wb['ConditionsPDE']
        wbConditionsPDEItems = wb['ConditionsPDEItems']
        wbConditions = wb['Conditions']
        wbMaterials = wb["materials"]
        wbPolygons = wb["polygons"]
        wbSheet = Modules.ManageFiles.ManageFiles.wbSheet(self, wb1, wb2, wb3, wb4, wb5,
        wb6, wb7, wb8, wbMatrixItems, wbConditionsPDE, wbConditionsPDEItems, wbConditions, wbPolygons, wbMaterials)

        #Cargar las figuras guardadas en el archivo Excel
        LoadExcel.loadExcelFigures(self, wbSheet, canvas)
        #Cargar las dimensiones de las matrices 
        LoadExcel.loadExcelMatrixDimensions(self, sheet, canvas, win)
        #Cargar los datos de los items del Coefficient PDE
        LoadExcel.loadExcelItemsData(self, wbMatrixItems)    
        #Cargar las coordenadas de los QCombobox del Coefficient PDE
        LoadExcel.loadExcelCoordinates(self, sheet)        
        #Cargar la lista de items activados del Coefficient PDE
        #LoadExcel.loadExcelItemsCoefficientPDE(self)
        #Cargar los datos de las matrices
        LoadExcel.loadExcelMatrixData(self, wbSheet)
        #Cargar los elementos de la matriz Conditions PDE del archivo Excel
        LoadExcel.loadExcelConditionsPDE(self, wbConditionsPDE)
        #Cargar los items activados de la clase Conditions PDE
        LoadExcel.loadExcelConditionsPDEItems(self, wbConditionsPDEItems)
        #Cargar los antiguos datos mostrados de cada coordenada del combobox
        LoadExcel.loadExcelCoordinateData(self)
        #Cargar los datos de la clase materials del archivo Excel
        LoadExcel.loadExcelMaterialsData(self, wbMaterials, material)
        #Cargar los datos de la clase conditions del archivo Excel
        LoadExcel.loadExcelConditionsData(self, wbConditions, condition)
        #Cargar la configuracion mas reciente del ModelWizard
        LoadExcel.loadExcelModelWizard(self, canvas)
        #Cargar la secuencia de paginas activadas en el archivo Excel
        LoadExcel.loadExcelSequenceTab(self, sheet, tabs)
        
       
        #Decirle al programa que no hay ediciones sen el archivo actual
        FileData.uncheckUpdateFile(self)
        self.actionSave_As.setEnabled(True)
        self.actionClose.setEnabled(True)



    def resetData(self, material, canvas):
        #Resetear los items de Coefficients PDE
        Reset.resetItemsCoefficientPDE(self)
        #Resetear coordenadas de Combobox y limpiar QlineEdits
        Reset.resetCoordinatesPDE(self)
        #Eliminar referencias al cualquier archivo
        Reset.resetUpdateFile(self)
        #Resetear configuracion de los items de Coefficient PDE
        Reset.resetItemsConfig(self)
        #Resetear la configuracion del ModelWizard
        Reset.resetModelWizard(self)
        #Resetear la configuracion de materials
        Reset.resetMaterials(self, material)
        #Resetear las figuras
        Reset.resetFigures(self, canvas)
        #Indicarle al programa que no hay ediciones en el archivo
        FileData.uncheckUpdateFile(self)



    def resetDataWithoutLoseFile(self):
        #Resetear los items de Coefficients PDE
        Reset.resetItemsCoefficientPDE(self)
        #Resetear coordenadas de Combobox y limpiar QlineEdits
        Reset.resetCoordinatesPDE(self)
        #Resetear configuracion de los items de Coefficient PDE
        Reset.resetItemsConfig(self)
        #Quitar la configuracion del Model Wizard actual
        Reset.removeModelWizard(self)

        if directory["dir"] != "":
         self.actionSaves.setEnabled(True)
         self.actionSave_As.setEnabled(True)
         self.actionClose.setEnabled(True)

        
class Update():
 #Funcion para actualizar los datos segun las coordenadas mas recientes
 def currentData(self, pos):
        if pos == 1:
            coordinates["coordinateDiffusion"] = [self.cmbRowDiffusionCoef.currentIndex(), self.cmbColumnDiffusionCoef.currentIndex()]
            floatMatrix = UpdateData.findCurrentInputMode(self, allNewMatrix.matrixCoefficientPDE[domains["domain"]][0])
            if floatMatrix == 0:
             UpdateData.setInputSingle(self, floatMatrix, allNewMatrix.matrixCoefficientPDE[domains["domain"]][0])
            if floatMatrix == 1:
             UpdateData.setInputDiagonal(self, floatMatrix, allNewMatrix.matrixCoefficientPDE[domains["domain"]][0])
            if floatMatrix == 2 or floatMatrix == 3:
             UpdateData.setInputSimmetryOrFull(self, floatMatrix, allNewMatrix.matrixCoefficientPDE[domains["domain"]][0])

        if pos == 2:
            coordinates["coordinateAbsorption"] = [self.cmbAbsorptionRow.currentIndex(), self.cmbAbsorptionColumn.currentIndex()]
            UpdateData.setCurrentSingleData(self, allNewMatrix.matrixCoefficientPDE[domains["domain"]][1][self.cmbAbsorptionRow.currentIndex()][self.cmbAbsorptionColumn.currentIndex()], self.lEditAbsorCoef)
        if pos == 3:
            coordinates["coordinateSource"] = [self.cmbSourceRow.currentIndex()]
            UpdateData.setCurrentSingleData(self, allNewMatrix.vectorCoefficientPDE[domains["domain"]][0][0][self.cmbSourceRow.currentIndex()], self.lEditSourceTerm)
        if pos == 4:
            coordinates["coordinateMass"] = [self.cmbMassCoefRow.currentIndex(), self.cmbMassCoefColumn.currentIndex()]
            UpdateData.setCurrentSingleData(self, allNewMatrix.matrixCoefficientPDE[domains["domain"]][2][self.cmbMassCoefRow.currentIndex()][self.cmbMassCoefColumn.currentIndex()], self.lEditMassCoef)
        if pos == 5:
            coordinates["coordinateDamMass"] = [self.cmbDamMassCoefRow.currentIndex(), self.cmbDamMassCoefColumn.currentIndex()]
            UpdateData.setCurrentSingleData(self, allNewMatrix.matrixCoefficientPDE[domains["domain"]][3][self.cmbDamMassCoefRow.currentIndex()][self.cmbDamMassCoefColumn.currentIndex()], self.lEditDamMassCoef)
        if pos == 6:
            coordinates["coordinateCFlux"] = [self.cmbCFluxRow.currentIndex(), self.cmbCFluxColumn.currentIndex()]
            UpdateData.setCurrentDoubleData(self, allNewMatrix.matrixCoefficientPDE[domains["domain"]][4][self.cmbCFluxRow.currentIndex()][self.cmbCFluxColumn.currentIndex()], self.lEditAlphaXCFlux, self.lEditAlphaCYFlux)
        if pos == 7:
            coordinates["coordinateConvection"] = [self.cmbConvectionRow.currentIndex(), self.cmbConvectionColumn.currentIndex()]
            UpdateData.setCurrentDoubleData(self, allNewMatrix.matrixCoefficientPDE[domains["domain"]][5][self.cmbConvectionRow.currentIndex()][self.cmbConvectionColumn.currentIndex()], self.lEditBetaXConvCoef, self.lEditBetaYConvCoef)
        if pos == 8: 
            coordinates["coordinateCSource"] = [self.cmbCSourceRow.currentIndex()]
            UpdateData.setCurrentDoubleData(self, allNewMatrix.vectorCoefficientPDE[domains["domain"]][1][0][self.cmbCSourceRow.currentIndex()], self.lEditGammaXCFluxSource, self.lEditGammaYCFluxSource)
            
#Funcion para actualizar la confiracion de los Combobox del Coefficient PDE
 def currentCoordinateMatrix(self, arrayComb):
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateDiffusion"], arrayComb[0])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateAbsorption"], arrayComb[1])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateSource"], arrayComb[2])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateMass"], arrayComb[3])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateDamMass"], arrayComb[4])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateCFlux"], arrayComb[5])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateConvection"], arrayComb[6])
    UpdateData.setCurrentCoordinateConfig(self, coordinates["coordinateCSource"], arrayComb[7])

        

        
        







        

        